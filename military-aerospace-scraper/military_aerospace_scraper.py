"""
Military Aerospace - Company Name & Email Scraper
==================================================
Extracts Company Name and Email from each company profile page.

Based on the HTML structure:
  - Name:  <h1 class="title-text" ...>Company Name</h1>
  - Email: <a href="mailto:someone@example.com" class="info" ...>

SETUP (run once):
-----------------
pip install playwright openpyxl
playwright install chromium

USAGE:
------
1. Paste your list of company URLs into the URLS list below.
2. Run:  python scrape_company_contacts.py
3. Results saved to:  military_companies_data.xlsx
"""

import time
import os
import re
from playwright.sync_api import sync_playwright
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ============================================================
#  PASTE YOUR COMPANY URLS HERE
# ============================================================
URLS = [
    "https://www.militaryaerospace.com/directory/services/distributors/company/14060405/aa-coatings-aerospace-thermal-spray-coating",
"https://www.militaryaerospace.com/directory/board-products/communications-network-controllers/company/14120504/a-flex",
"https://www.militaryaerospace.com/directory/company/14121203/a2e-ltd",
"https://www.militaryaerospace.com/directory/company/14118338/aaa-oem-odm-ltd",


]

OUTPUT_FILE    = "military_companies_data.xlsx"
MAX_RETRIES    = 3          # retries per URL on timeout/error
PAGE_TIMEOUT   = 30000      # ms — page load timeout (30s)
WAIT_AFTER     = 1.5        # seconds to pause between pages
HEADLESS       = True       # True = silent, False = see browser


# ============================================================
#  EXTRACTION LOGIC
# ============================================================
def extract_info(page) -> tuple[str, str]:
    """Extract company name and email from a loaded page."""

    name  = "Not Found"
    email = "Not Found"

    # --- Company Name ---
    for selector in ["h1.title-text", "h1"]:
        try:
            el = page.query_selector(selector)
            if el:
                t = el.inner_text().strip()
                if t:
                    name = t
                    break
        except Exception:
            pass

    # --- Email ---
    try:
        el = page.query_selector('a[href^="mailto:"]')
        if el:
            href = el.get_attribute("href") or ""
            email = href.replace("mailto:", "").strip()
            email = email.split("?")[0].strip()
    except Exception:
        pass

    # Fallback: scan all text for email pattern
    if email == "Not Found":
        try:
            body_text = page.inner_text("body")
            matches = re.findall(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}", body_text)
            if matches:
                bad = {"sentry", "example", "domain", "yourdomain", "noreply",
                       "privacy", "legal", "support@playwright"}
                filtered = [m for m in matches if not any(b in m.lower() for b in bad)]
                if filtered:
                    email = filtered[0]
        except Exception:
            pass

    return name, email


# ============================================================
#  EXCEL HELPERS
# ============================================================
def make_styles():
    thin = Side(style="thin", color="CCCCCC")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    return {
        "bdr":       bdr,
        "hdr_font":  Font(name="Arial", bold=True, color="FFFFFF", size=11),
        "hdr_fill":  PatternFill("solid", start_color="1A3C6E"),
        "hdr_align": Alignment(horizontal="center", vertical="center"),
        "fill_odd":  PatternFill("solid", start_color="EEF4FB"),
        "fill_even": PatternFill("solid", start_color="FFFFFF"),
        "num_font":  Font(name="Arial", size=10, color="666666"),
        "name_font": Font(name="Arial", size=10),
        "link_font": Font(name="Arial", size=10, color="1155CC", underline="single"),
        "ok_font":   Font(name="Arial", size=10, color="0A6B0A"),   # green  – email found
        "err_font":  Font(name="Arial", size=10, color="CC0000", italic=True),  # red
        "bold_font": Font(name="Arial", bold=True, size=10),
    }


def init_workbook(filename: str):
    """Create a fresh workbook with header row."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Military Contacts"
    s = make_styles()

    for col, h in enumerate(["#", "Company Name", "Email", "Status", "Source URL"], 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = s["hdr_font"]; c.fill = s["hdr_fill"]
        c.alignment = s["hdr_align"]; c.border = s["bdr"]
    ws.row_dimensions[1].height = 26

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 90
    ws.freeze_panes = "A2"
    wb.save(filename)


def append_row(filename: str, row_num: int, name: str, email: str,
               status: str, url: str):
    """Append one result row to the existing workbook."""
    wb = load_workbook(filename)
    ws = wb.active
    s  = make_styles()

    r    = row_num + 1           # row 1 = header
    fill = s["fill_odd"] if row_num % 2 == 1 else s["fill_even"]

    found = email not in ("Not Found", "Error")

    c1 = ws.cell(row=r, column=1, value=row_num)
    c1.font = s["num_font"]; c1.fill = fill
    c1.alignment = Alignment(horizontal="center", vertical="center")
    c1.border = s["bdr"]

    c2 = ws.cell(row=r, column=2, value=name)
    c2.font = s["name_font"]; c2.fill = fill
    c2.alignment = Alignment(vertical="center"); c2.border = s["bdr"]

    c3 = ws.cell(row=r, column=3, value=email)
    c3.font = s["ok_font"] if found else s["err_font"]
    c3.fill = fill
    c3.alignment = Alignment(vertical="center"); c3.border = s["bdr"]

    c4 = ws.cell(row=r, column=4, value=status)
    c4.font = s["ok_font"] if status == "OK" else s["err_font"]
    c4.fill = fill
    c4.alignment = Alignment(horizontal="center", vertical="center")
    c4.border = s["bdr"]

    c5 = ws.cell(row=r, column=5, value=url)
    c5.hyperlink = url
    c5.font = s["link_font"]; c5.fill = fill
    c5.alignment = Alignment(vertical="center"); c5.border = s["bdr"]

    ws.row_dimensions[r].height = 16
    wb.save(filename)


def write_summary(filename: str, total: int, found: int, errors: int):
    wb = load_workbook(filename)
    ws = wb.active
    s  = make_styles()
    tr = total + 2
    ws.cell(row=tr, column=1, value="Total:").font  = s["bold_font"]
    ws.cell(row=tr, column=2, value=total).font     = s["bold_font"]
    ws.cell(row=tr, column=3,
            value=f"{found} emails found / {errors} errors").font = s["bold_font"]
    wb.save(filename)


# ============================================================
#  MAIN SCRAPER
# ============================================================
def main():
    if not URLS:
        print("⚠  URLS list is empty. Add URLs to the script and re-run.")
        return

    # Load already-processed URLs so we can resume if script was restarted
    done_urls: set[str] = set()
    start_row = 1
    if os.path.exists(OUTPUT_FILE):
        try:
            wb = load_workbook(OUTPUT_FILE)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[4]:                   # column E = URL
                    done_urls.add(str(row[4]))
            start_row = len(done_urls) + 1
            print(f"📂 Resuming — {len(done_urls)} URLs already done, "
                  f"continuing from #{start_row}")
        except Exception:
            pass
    else:
        init_workbook(OUTPUT_FILE)

    pending = [u for u in URLS if u not in done_urls]
    total   = len(URLS)
    found_count  = start_row - 1   # already found in previous run
    error_count  = 0

    print(f"\n{'='*60}")
    print(f"  Military Aerospace — Company Contact Scraper")
    print(f"{'='*60}")
    print(f"  Total URLs   : {total}")
    print(f"  To process   : {len(pending)}")
    print(f"  Output file  : {OUTPUT_FILE}")
    print(f"  Headless     : {HEADLESS}")
    print(f"{'='*60}\n")

    row_num = start_row

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=HEADLESS)
        context = browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            )
        )
        # Block images / fonts — makes pages load much faster
        context.route(
            "**/*.{png,jpg,jpeg,gif,webp,svg,ico,woff,woff2,ttf,eot,mp4,mp3}",
            lambda r: r.abort()
        )
        page = context.new_page()

        for url in pending:
            global_idx = URLS.index(url) + 1
            print(f"[{global_idx}/{total}] {url[:80]}...")

            name   = "Not Found"
            email  = "Not Found"
            status = "Error"

            for attempt in range(1, MAX_RETRIES + 1):
                try:
                    page.goto(url, wait_until="domcontentloaded",
                              timeout=PAGE_TIMEOUT)

                    # Wait for the main content block
                    try:
                        page.wait_for_selector(
                            "h1.title-text, h1",
                            timeout=10000
                        )
                    except Exception:
                        pass   # proceed anyway

                    time.sleep(1.2)   # let JS render
                    name, email = extract_info(page)
                    status = "OK"
                    break             # success — stop retrying

                except Exception as e:
                    print(f"  ⚠ Error on attempt {attempt}/{MAX_RETRIES}: {e}")
                    if attempt < MAX_RETRIES:
                        time.sleep(3 * attempt)   # back-off
                    else:
                        status = "Error"

            # Count stats
            if email not in ("Not Found",):
                found_count += 1
            if status != "OK":
                error_count += 1

            # Write row immediately — so progress is never lost
            append_row(OUTPUT_FILE, row_num, name, email, status, url)
            row_num += 1

            icon = "✔" if status == "OK" else "✘"
            print(f"  {icon}  Name  : {name}")
            print(f"     Email : {email}  [{status}]\n")

            time.sleep(WAIT_AFTER)

        browser.close()

    write_summary(OUTPUT_FILE, total, found_count, error_count)

    print(f"\n{'='*60}")
    print(f"✅  Done!")
    print(f"   Total processed : {total}")
    print(f"   Emails found    : {found_count}")
    print(f"   Errors/Timeouts : {error_count}")
    print(f"   Output file     : {OUTPUT_FILE}")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()