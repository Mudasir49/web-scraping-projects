"""
==============================================================
  LEADS ENRICHER — Fast Version
  Website Scraper + Hunter.io API
  
  Reads:  tax_leads_results.xlsx
  Saves:  tax_leads_enriched.csv  (live, open anytime)
  
  New in this version:
  - 5 parallel workers (5x faster)
  - 8s timeout (faster skip of dead sites)
  - Facebook & YouTube extracted if no email found
  - All existing features kept
==============================================================
"""

import requests
import re
import time
import os
import csv
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from concurrent.futures import ThreadPoolExecutor, as_completed

# ============================================================
HUNTER_API_KEY   = "YOUR_HUNTER_IO_API_KEY_HERE"
INPUT_FILE       = "tax_leads_results.xlsx"
OUTPUT_FILE      = "tax_leads_enriched.csv"
HUNTER_LIMIT     = 25
PARALLEL_WORKERS = 5
TIMEOUT          = 8
# ============================================================

BROWSER_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Connection": "keep-alive",
}

SKIP_EMAIL_WORDS = [
    "example", "domain", "test", "png", "jpg", "gif", "wix",
    "wordpress", "sentry", "schema", "jquery", "bootstrap",
    "fontawesome", "google", "facebook", "twitter", "linkedin",
    "instagram", "youtube", "privacy", "terms", "noreply",
    "no-reply", "placeholder", "youremail", "email@"
]

OWNER_TITLES = [
    "owner", "founder", "co-founder", "president", "principal",
    "managing partner", "ceo", "chief executive", "director",
    "managing director", "partner", "cpa", "enrolled agent",
    "tax attorney", "chief financial"
]

CSV_HEADERS = [
    # Original fields
    "Company Name", "Full Address", "City", "State", "State Code",
    "Zip Code", "Phone", "Website", "Business Category",
    "Search Keyword", "Search State",
    "Bing Maps Link", "Google Search Link", "API Source Link",
    "Date Collected",
    # Enriched fields
    "Company Email",
    "Owner Name",
    "Owner Phone",
    "Owner Email",
    "Facebook Page",
    "YouTube Channel",
    "Enrichment Source",
    "Data Found On URL",
]


# ============================================================
#   HELPERS
# ============================================================

def clean_email(email):
    email = email.lower().strip()
    if any(skip in email for skip in SKIP_EMAIL_WORDS):
        return None
    if not re.match(r"^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$", email):
        return None
    if len(email) > 100:
        return None
    return email


def extract_phones(text):
    phones = re.findall(
        r"(?:\+1[\s\-.]?)?\(?\d{3}\)?[\s\-.]?\d{3}[\s\-.]?\d{4}", text
    )
    clean = []
    for p in phones:
        digits = re.sub(r"\D", "", p)
        if len(digits) in [10, 11]:
            clean.append(p.strip())
    return list(set(clean))


def get_domain(url):
    if not url:
        return ""
    url = url.replace("https://", "").replace("http://", "").replace("www.", "")
    return url.split("/")[0].strip()


def extract_social_links(soup, html_text):
    """Extract Facebook and YouTube links from page."""
    facebook  = ""
    youtube   = ""

    # Facebook patterns
    fb_patterns = [
        r'https?://(?:www\.)?facebook\.com/(?!sharer|share|login|home)([A-Za-z0-9._\-/]+)',
        r'https?://(?:www\.)?fb\.com/([A-Za-z0-9._\-/]+)',
    ]
    for pattern in fb_patterns:
        matches = re.findall(pattern, html_text)
        for m in matches:
            # Skip generic/share URLs
            if not any(skip in m.lower() for skip in ["sharer", "share", "login", "dialog", "plugins", "photo"]):
                facebook = f"https://facebook.com/{m.rstrip('/')}"
                break
        if facebook:
            break

    # YouTube patterns
    yt_patterns = [
        r'https?://(?:www\.)?youtube\.com/(?:channel|c|user|@)([A-Za-z0-9._\-/]+)',
        r'https?://(?:www\.)?youtube\.com/([A-Za-z0-9._\-]+)',
    ]
    for pattern in yt_patterns:
        matches = re.findall(pattern, html_text)
        for m in matches:
            if not any(skip in m.lower() for skip in ["watch", "embed", "playlist", "shorts", "results"]):
                youtube = f"https://youtube.com/{m.rstrip('/')}"
                break
        if youtube:
            break

    return facebook, youtube


# ============================================================
#   WEBSITE SCRAPER — runs per company
# ============================================================

def scrape_company(row_data):
    """
    Full enrichment for one company.
    Returns dict with all enriched fields.
    Designed to run in parallel.
    """
    company_name = row_data.get("Company Name", "")
    website      = row_data.get("Website", "")

    result = {
        "Company Email": "",
        "Owner Name":    "",
        "Owner Phone":   "",
        "Owner Email":   "",
        "Facebook Page": "",
        "YouTube Channel": "",
        "Enrichment Source": "not_found",
        "Data Found On URL": "",
    }

    # ── No website — instant skip ──
    if not website or str(website).strip() in ["", "None", "N/A", "none"]:
        result["Enrichment Source"] = "no_website"
        return result

    if not website.startswith("http"):
        website = "https://" + website

    base_url = website.rstrip("/")

    pages_to_try = [
        base_url,
        base_url + "/contact",
        base_url + "/contact-us",
        base_url + "/about",
        base_url + "/about-us",
        base_url + "/team",
        base_url + "/our-team",
        base_url + "/staff",
    ]

    company_email       = ""
    company_email_page  = ""
    owner_name          = ""
    owner_name_page     = ""
    owner_phone         = ""
    owner_phone_page    = ""
    owner_email         = ""
    owner_email_page    = ""
    facebook            = ""
    youtube             = ""
    all_emails_by_page  = {}
    pages_scraped       = 0

    for (page_url, page_label) in [(p, "") for p in pages_to_try]:
        if pages_scraped >= 3:
            break
        try:
            r = requests.get(
                page_url,
                headers=BROWSER_HEADERS,
                timeout=TIMEOUT,
                allow_redirects=True
            )
            if r.status_code != 200:
                continue

            actual_url    = r.url
            pages_scraped += 1
            html_text     = r.text

            soup = BeautifulSoup(html_text, "lxml")
            for tag in soup(["script", "style", "noscript", "meta"]):
                tag.decompose()
            page_text = soup.get_text(" ", strip=True)

            # ── Emails ──
            raw_emails  = re.findall(
                r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}",
                html_text
            )
            page_emails = [clean_email(e) for e in raw_emails]
            page_emails = [e for e in page_emails if e]

            if page_emails:
                all_emails_by_page[actual_url] = page_emails
                if not company_email:
                    company_email      = page_emails[0]
                    company_email_page = actual_url

            # ── Phones ──
            if not owner_phone:
                phones = extract_phones(page_text)
                if phones:
                    owner_phone      = phones[0]
                    owner_phone_page = actual_url

            # ── Owner name ──
            if not owner_name:
                for title in OWNER_TITLES:
                    patterns = [
                        rf"([A-Z][a-z]{{2,}} [A-Z][a-z]{{2,}}(?:\s[A-Z][a-z]{{2,}})?)[,\s]{{0,15}}{title}",
                        rf"{title}[:\s,]{{1,10}}([A-Z][a-z]{{2,}} [A-Z][a-z]{{2,}}(?:\s[A-Z][a-z]{{2,}})?)",
                    ]
                    for pattern in patterns:
                        matches = re.findall(pattern, page_text, re.IGNORECASE)
                        if matches:
                            candidate = matches[0].strip()
                            if 2 <= len(candidate.split()) <= 4 and len(candidate) < 50:
                                owner_name      = candidate
                                owner_name_page = actual_url
                                break
                    if owner_name:
                        break

                if not owner_name:
                    for tag in soup.find_all(["h2", "h3", "h4", "p", "div", "span"]):
                        tag_text = tag.get_text(" ", strip=True)
                        for title in OWNER_TITLES:
                            if title.lower() in tag_text.lower() and len(tag_text) < 120:
                                m = re.search(r"([A-Z][a-z]{2,} [A-Z][a-z]{2,})", tag_text)
                                if m:
                                    owner_name      = m.group(1)
                                    owner_name_page = actual_url
                                    break
                        if owner_name:
                            break

            # ── Social links (check every page until both found) ──
            if not facebook or not youtube:
                fb, yt = extract_social_links(soup, html_text)
                if fb and not facebook:
                    facebook = fb
                if yt and not youtube:
                    youtube = yt

        except requests.exceptions.Timeout:
            continue
        except Exception:
            continue

    # ── Match owner email ──
    if owner_name:
        name_parts = owner_name.lower().split()
        for page_url, emails in all_emails_by_page.items():
            for email in emails:
                local = email.split("@")[0].lower()
                if any(part in local for part in name_parts if len(part) > 2):
                    owner_email      = email
                    owner_email_page = page_url
                    break
            if owner_email:
                break

    # Company email != owner email
    if company_email == owner_email and owner_email:
        for page_url, emails in all_emails_by_page.items():
            for email in emails:
                if email != owner_email:
                    company_email      = email
                    company_email_page = page_url
                    break
            if company_email != owner_email:
                break

    # ── Only add social if NO email found ──
    email_found = bool(company_email or owner_email)
    if email_found:
        facebook = ""
        youtube  = ""

    # ── Build source pages string ──
    source_pages = []
    if company_email_page:
        source_pages.append(f"company_email:{company_email_page}")
    if owner_name_page:
        source_pages.append(f"owner_name:{owner_name_page}")
    if owner_phone_page:
        source_pages.append(f"owner_phone:{owner_phone_page}")
    if owner_email_page:
        source_pages.append(f"owner_email:{owner_email_page}")

    data_found_on = " | ".join(source_pages)

    has_data = any([company_email, owner_name, owner_phone, owner_email, facebook, youtube])

    result.update({
        "Company Email":    company_email,
        "Owner Name":       owner_name,
        "Owner Phone":      owner_phone,
        "Owner Email":      owner_email,
        "Facebook Page":    facebook,
        "YouTube Channel":  youtube,
        "Enrichment Source": "website" if has_data else "not_found",
        "Data Found On URL": data_found_on,
    })
    return result


# ============================================================
#   HUNTER.IO API
# ============================================================

def hunter_domain_search(domain):
    try:
        r = requests.get(
            "https://api.hunter.io/v2/domain-search",
            params={
                "domain":  domain,
                "api_key": HUNTER_API_KEY,
                "limit":   10,
            },
            timeout=10
        )

        if r.status_code == 200:
            data          = r.json().get("data", {})
            emails        = data.get("emails", [])
            company_email = ""
            owner_name    = ""
            owner_email   = ""
            hunter_url    = f"https://hunter.io/domain-search/{domain}"

            for entry in emails:
                email     = entry.get("value", "")
                position  = (entry.get("position") or "").lower()
                fname     = entry.get("first_name") or ""
                lname     = entry.get("last_name")  or ""
                full_name = f"{fname} {lname}".strip()

                if not company_email and email:
                    company_email = email
                if any(t in position for t in OWNER_TITLES):
                    if not owner_name  and full_name:
                        owner_name  = full_name
                    if not owner_email and email:
                        owner_email = email

            result = {
                "Company Email": company_email,
                "Owner Name":    owner_name,
                "Owner Email":   owner_email,
            }
            data_found_on = hunter_url if any(result.values()) else ""
            return result, "hunter", data_found_on

        elif r.status_code == 429:
            print("\n  ⚠️  Hunter.io rate limit — waiting 60s...")
            time.sleep(60)
            return {}, "hunter_limit", ""
        else:
            return {}, "hunter_error", ""

    except Exception:
        return {}, "hunter_exception", ""


# ============================================================
#   RESUME SUPPORT
# ============================================================

def load_done_companies():
    done = set()
    if not os.path.exists(OUTPUT_FILE):
        return done
    try:
        with open(OUTPUT_FILE, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                if row.get("Company Name"):
                    done.add(row["Company Name"].strip().lower())
    except:
        pass
    return done


# ============================================================
#   MAIN
# ============================================================

def main():
    print("=" * 65)
    print("   LEADS ENRICHER — Fast Version")
    print(f"   {PARALLEL_WORKERS} parallel workers | {TIMEOUT}s timeout")
    print("=" * 65)

    # ── Load input ──
    if not os.path.exists(INPUT_FILE):
        print(f"\n  ❌ Not found: {INPUT_FILE}")
        return

    print(f"\n  📂 Loading {INPUT_FILE}...")
    wb      = load_workbook(INPUT_FILE, read_only=True)
    ws      = wb["All Leads"]
    headers = [cell.value for cell in ws[1]]
    rows    = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    print(f"  ✅ {len(rows)} companies loaded")

    def col(name):
        try:    return headers.index(name)
        except: return -1

    # Build row dicts
    all_row_dicts = []
    for row in rows:
        d = {h: (row[i] if i >= 0 and i < len(row) else "") for i, h in enumerate(headers)}
        all_row_dicts.append(d)

    # Resume
    done_companies = load_done_companies()
    if done_companies:
        print(f"  ♻️  Resuming — {len(done_companies)} already enriched")

    # Filter pending
    pending_rows = [
        r for r in all_row_dicts
        if r.get("Company Name") and
        str(r.get("Company Name", "")).strip().lower() not in done_companies
    ]

    total        = len(all_row_dicts)
    pending      = len(pending_rows)
    file_exists  = os.path.exists(OUTPUT_FILE) and os.path.getsize(OUTPUT_FILE) > 0
    hunter_used  = 0
    found_email  = 0
    found_owner  = 0
    found_social = 0
    completed    = 0
    start_time   = time.time()

    print(f"\n  Total    : {total}")
    print(f"  Pending  : {pending}")
    print(f"  Workers  : {PARALLEL_WORKERS} parallel")
    print(f"  Hunter   : {HUNTER_LIMIT} free/month")
    print(f"\n  Open {OUTPUT_FILE} anytime — updates live!\n")

    with open(OUTPUT_FILE, mode="a", newline="", encoding="utf-8") as csvfile:
        writer = csv.writer(csvfile)

        if not file_exists:
            writer.writerow(CSV_HEADERS)
            csvfile.flush()

        # Process in parallel batches
        for i in range(0, len(pending_rows), PARALLEL_WORKERS):
            batch = pending_rows[i:i + PARALLEL_WORKERS]

            with ThreadPoolExecutor(max_workers=PARALLEL_WORKERS) as executor:
                futures = {
                    executor.submit(scrape_company, row_dict): row_dict
                    for row_dict in batch
                }

                for future in as_completed(futures):
                    row_dict   = futures[future]
                    enrichment = future.result()
                    completed += 1

                    company_name = row_dict.get("Company Name", "")
                    website      = row_dict.get("Website", "")

                    # ── Hunter.io for gaps (sequential — API rate limit) ──
                    needs_hunter = (
                        not enrichment["Company Email"] and
                        not enrichment["Owner Email"] and
                        enrichment["Enrichment Source"] not in ["no_website"] and
                        website and
                        hunter_used < HUNTER_LIMIT
                    )
                    if needs_hunter:
                        domain = get_domain(str(website))
                        if domain:
                            h_data, h_src, h_url = hunter_domain_search(domain)
                            hunter_used += 1
                            for k, v in h_data.items():
                                if v and not enrichment.get(k):
                                    enrichment[k] = v
                            if any(h_data.values()):
                                src = enrichment["Enrichment Source"]
                                enrichment["Enrichment Source"] = "hunter" if src in ["not_found", ""] else src + "+hunter"
                                fou = enrichment["Data Found On URL"]
                                enrichment["Data Found On URL"] = (fou + " | " + h_url).strip(" | ") if fou else h_url
                            time.sleep(1.5)

                    # ── Write row immediately ──
                    csv_row = []
                    for header in CSV_HEADERS:
                        if header in enrichment:
                            csv_row.append(enrichment.get(header, ""))
                        else:
                            csv_row.append(row_dict.get(header, ""))
                    writer.writerow(csv_row)
                    csvfile.flush()

                    # ── Stats ──
                    has_email  = bool(enrichment["Company Email"] or enrichment["Owner Email"])
                    has_owner  = bool(enrichment["Owner Name"])
                    has_social = bool(enrichment["Facebook Page"] or enrichment["YouTube Channel"])
                    if has_email:  found_email  += 1
                    if has_owner:  found_owner  += 1
                    if has_social: found_social += 1

                    elapsed = time.time() - start_time
                    speed   = completed / elapsed if elapsed > 0 else 1
                    eta_min = (pending - completed) / speed / 60 if speed > 0 else 0

                    icons = ""
                    icons += "📧" if has_email  else "  "
                    icons += "👤" if has_owner  else "  "
                    icons += "📘" if enrichment["Facebook Page"]   else "  "
                    icons += "▶️ " if enrichment["YouTube Channel"] else "  "

                    print(
                        f"  [{completed:>4}/{pending}] "
                        f"{str(company_name)[:35]:<35} "
                        f"{icons} "
                        f"src={enrichment['Enrichment Source']:<12} "
                        f"ETA:{eta_min:>5.1f}min"
                    )

                    if hunter_used >= HUNTER_LIMIT:
                        print(f"\n  ℹ️  Hunter.io limit reached — website only\n")

    elapsed_total = (time.time() - start_time) / 60
    print(f"\n{'=' * 65}")
    print(f"  DONE!")
    print(f"  Processed      : {completed}")
    print(f"  Emails found   : {found_email}")
    print(f"  Owners found   : {found_owner}")
    print(f"  Social found   : {found_social} (no email companies)")
    print(f"  Hunter used    : {hunter_used}/{HUNTER_LIMIT}")
    print(f"  Time taken     : {elapsed_total:.1f} minutes")
    print(f"  Output         : {OUTPUT_FILE}")
    print(f"{'=' * 65}\n")


if __name__ == "__main__":
    main()